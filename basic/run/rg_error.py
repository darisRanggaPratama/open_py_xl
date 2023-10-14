import openpyxl as pyxl

# Buka file Workbook
path = input('Error Values in Worksheet\nFile Name: ')
workbook = pyxl.load_workbook(path)

# Pilih Worksheet, misal: "GAJI"
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi variabel untuk menyimpan alamat sel dan nilainya
hasil_pencarian = []

# Loop melalui sel, misal: A2 hingga F46
for row in worksheet.iter_rows(min_row=2, max_row=46, min_col=1, max_col=6):
    for cell in row:
        if isinstance(cell.value, str):
            # Cek panjang nilai cell
            if len(cell.value) > 0:
                # Cek karakter dicari
                if cell.value[0] == '#':
                    # Tambahkan hasil ke list
                    hasil_pencarian.append((cell.coordinate, cell.value))

# Cetak hasil pencarian
print('\nHasil pencarian nilai\nError:\n\n No Cell Value')
x = 0
for alamat, nilai in hasil_pencarian:
    x = x + 1
    print(f' {x}   {alamat}  {nilai}')

print('\n===End Searching===')
# Tutup file Workbook
workbook.close()
