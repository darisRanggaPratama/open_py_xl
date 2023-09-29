import openpyxl as oxl

wbook = oxl.load_workbook('sample.xlsx')
print(type(wbook))
print(wbook.sheetnames)
wsheet = wbook['GAJI']
print(wsheet)
print(type(wsheet))
print(wsheet.title)
actSheet = wbook.active
print(actSheet)

# Read workbook, worksheet, active sheet:

# 1. import openpyxl as oxl: Mengimpor pustaka openpyxl dan memberikan alias kepadanya sebagai oxl.

# 2. wbook = oxl.load_workbook('sample.xlsx'): Akses file Excel 'sample.xlsx' ke dalam variabel
# wbook menggunakan fungsi load_workbook dari pustaka openpyxl.

# 3. print(type(wbook)): Tampilkan isi variable/ objek wbook. Hasilnya akan mencetak
# tipe objek: <class 'openpyxl.workbook.workbook.Workbook'>.

# 4. print(wbook.sheetnames): Tampilkan nama-nama worksheet yang ada dalam buku kerja Excel.
# Fungsi sheetnames mengembalikan daftar nama worksheet di dalam workbook.

# 5. wsheet = wbook['GAJI']: Tampilkan worksheet dengan nama 'GAJI' dari workbook Excel
# dan menyimpannya dalam variabel wsheet.

# 6. print(wsheet): Baris ini mencetak objek worksheet yang disimpan dalam variabel wsheet.

# 7. print(type(wsheet)): Ini mencetak tipe objek yang disimpan dalam variabel wsheet.
# Tampilkan tipe objek: <class 'openpyxl.worksheet.worksheet.Worksheet'>.

# 8. print(wsheet.title): Ini mencetak judul (nama) lembar yang disimpan dalam variabel wsheet.

# 9. actSheet = wbook.active: Mengambil worksheet aktif (worksheet yang sedang aktif digunakan)
# dalam workbook Excel dan menyimpannya dalam variabel actSheet.

# 10. print(actSheet): Baris ini mencetak objek lembar aktif yang disimpan dalam variabel actSheet.

# Kesimpulan: Secara keseluruhan, kode tersebut digunakan untuk membuka file Excel,
# mengakses worksheet di dalamnya, dan mencetak beberapa informasi
# terkait dengan workbook dan worksheet tersebut.
