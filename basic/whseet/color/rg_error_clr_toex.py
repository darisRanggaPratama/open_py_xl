import openpyxl as pyxl
from openpyxl.styles import PatternFill

# Buka file Excel: *.xlsx
path = input('Error Values in Worksheet\nFile Name: ')
workbook = pyxl.load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Warna latar belakang
colored = PatternFill(start_color='EDBB99', end_color='EDBB99', fill_type='solid')

# Loop melalui sel di Worksheet target, misal dari A2 sampai F46
for row in worksheet.iter_rows(min_row=2, max_row=46, min_col=1, max_col=6):
    for cell in row:
        # Periksa apakah nilai Error
        if isinstance(cell.value, str):
            # Cek panjang nilai cell
            if len(cell.value) > 0:
                # Cek karakter dicari
                if cell.value.startswith('#'):
                    cell.fill = colored
                    # Tambahkan hasil ke list
                    result.append((cell.coordinate, cell.value))

# Buat worksheet "ERR" jika belum ada
if 'ERR' not in workbook.sheetnames:
    workbook.create_sheet('ERR')

# Pilih worksheet "ERR"
sheetErr = workbook['ERR']

# Judul: ERROR
sheetErr.cell(row=1, column=2, value='No; Cell; ERROR')

# Tulis hasil ke dalam Worksheet "ERR" di cell B2
for i, (coordinate, value) in enumerate(result):
    sheetErr.cell(row=i + 2, column=2, value=f"{i + 1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file Excel: *.xlsx
workbook.save('sample.xlsx')
