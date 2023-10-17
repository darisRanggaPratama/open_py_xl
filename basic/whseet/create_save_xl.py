import openpyxl as pyxl

wbook = pyxl.Workbook()
print(f'sheets: {wbook.sheetnames}')
wsheet = wbook.active
print(f'title: {wsheet.title}')
wsheet.title = 'data-baru'
print(f'Nama sheet: {wbook.sheetnames}')
