from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment

# Buka file Excel: *.xlsx
path = input('Negative Values in Worksheet\nFile Name: ')
workbook = load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Warna latar belakang
colored = PatternFill(start_color='D2B4DE', end_color='D2B4DE', fill_type='solid')

# Font italic
italicFont = Font(italic=True)

# Loop melalui sel di Worksheet target, misal dari A3 sampai E47
for row in worksheet.iter_rows(min_row=3, max_row=47, min_col=1, max_col=5):
    for cell in row:
        # Periksa apakah nilai Negative
        if isinstance(cell.value, (int, float)):
            # Cek nilai sel negative
            negative = cell.value < 0
            if negative is True:
                commented = Comment(f"NEGATIVE\n{cell.coordinate}: {cell.value}", "Author")
                cell.comment = None
                cell.comment = commented
                cell.font = italicFont
                cell.fill = colored
                # Tambahkan hasil ke list
                result.append((cell.coordinate, cell.value))

# Buat worksheet "NEG" jika belum ada
if 'NEG' not in workbook.sheetnames:
    workbook.create_sheet('NEG')

# Pilih worksheet "NEG"
sheetNeg = workbook['NEG']

# Judul: NEGATIVE
sheetNeg.cell(row=1, column=2, value='No; Cell; NEGATIVE')

# Tulis hasil ke dalam Worksheet "NEG" di cell B2
for i, (coordinate, value) in enumerate(result):
    sheetNeg.cell(row=i + 2, column=2, value=f"{i + 1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file Excel: *.xlsx
workbook.save(path)
