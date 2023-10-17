import openpyxl as pyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Buka file Excel: *.xlsx
path = input('String Values in Worksheet\nFile Name: ')
workbook = pyxl.load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Warna latar belakang
colored = PatternFill(start_color='D5DBDB', end_color='D5DBDB', fill_type='solid')

# Loop melalui sel di Worksheet target, misal dari A2 sampai F46
for row in worksheet.iter_rows(min_row=4, max_row=1145, min_col=5, max_col=21):
    for cell in row:
        # Periksa apakah nilai Error
        if isinstance(cell.value, str):
            # Cek panjang nilai cell
            if len(cell.value) > 0:
                # Cek karakter dicari
                if not cell.value.startswith('#'):
                    commented = Comment(f"TEXT\n{cell.coordinate}: {cell.value}", "Author")
                    cell.comment = None
                    cell.comment = commented
                    cell.fill = colored
                    # Tambahkan hasil ke list
                    result.append((cell.coordinate, cell.value))

# Buat worksheet "STR" jika belum ada
if 'STR' not in workbook.sheetnames:
    workbook.create_sheet('STR')

# Pilih worksheet "STR"
sheetErr = workbook['STR']

# Judul: STRING
sheetErr.cell(row=1, column=2, value='No; Cell; STRING')

# Tulis hasil ke dalam Worksheet "STR" di cell B2
for i, (coordinate, value) in enumerate(result):
    sheetErr.cell(row=i + 2, column=2, value=f"{i + 1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file Excel: *.xlsx
workbook.save(path)
