from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Buka file Excel: *.xlsx
path = input('Null Values in Worksheet\nFile Name: ')
workbook = load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Warna latar belakang
colored = PatternFill(start_color='85C1E9', end_color='85C1E9', fill_type='solid')

# Loop melalui sel di Worksheet target, misal dari A2 sampai F46
for row in worksheet.iter_rows(min_row=2, max_row=46, min_col=1, max_col=6):
    for cell in row:
        # Periksa apakah nilai Blank/ Null
        if cell.value is None:
            commented = Comment(f"NULL\n{cell.coordinate}: {cell.value}", "Author")
            cell.comment = None
            cell.comment = commented
            cell.fill = colored            
            # Tambahkan hasil ke list
            result.append((cell.coordinate, cell.value))

# Buat worksheet "NULL" jika belum ada
if 'NUL' not in workbook.sheetnames:
    workbook.create_sheet('NUL')

# Pilih worksheet "NULL"
sheetNul = workbook['NUL']

# Judul: NULL
sheetNul.cell(row=1, column=2, value='No; Cell; NULL')

# Tulis hasil ke dalam Worksheet "NULL" di cell B2
for i, (coordinate, value) in enumerate(result):
    sheetNul.cell(row=i + 2, column=2, value=f"{i + 1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file Excel: *.xlsx
workbook.save(path)
