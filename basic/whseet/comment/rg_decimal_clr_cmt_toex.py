import openpyxl as pyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Buka file Excel: *.xlsx
path = input('Decimal Values in Worksheet\nFile Name: ')
workbook = pyxl.load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Warna latar belakang
colored = PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')

# Loop melalui sel di Worksheet target, misal dari E4 sampai U1145
for row in worksheet.iter_rows(min_row=4, max_row=1145, min_col=5, max_col=21):
    for cell in row:
        # Periksa apakah nilai sel adalah Decimal/Float/Double
        if isinstance(cell.value, float):
            commented = Comment(f"DECIMAL\n{cell.coordinate}: {cell.value}", "Author")
            cell.comment = None
            cell.comment = commented
            cell.fill = colored
            result.append((cell.coordinate, cell.value))

# Buat worksheet "DEC" jika belum ada
if 'DEC' not in workbook.sheetnames:
    workbook.create_sheet('DEC')

# Pilih worksheet "DEC"
sheetDec = workbook['DEC']

# Judul: Desimal
sheetDec.cell(row=1, column=2, value='No; Cell; DECIMAL')

# Tulis hasil ke dalam Worksheet "DEC" di cell B2
for i, (coordinate, value) in enumerate(result):
    sheetDec.cell(row=i + 2, column=2, value=f"{i + 1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file.xlsx
workbook.save(path)
