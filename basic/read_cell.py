import openpyxl as oxl
from openpyxl.utils import *

wbook = oxl.load_workbook('run/sample.xlsx')
print(type(wbook))
print(wbook.sheetnames)
wsheet = wbook['GAJI']
cell = wsheet['A9']
print(
    f"Row: {cell.row} "
    f"Column: {cell.column} / {get_column_letter(cell.column)} "
    f"Position: {cell.coordinate} Value: {cell.value}")
