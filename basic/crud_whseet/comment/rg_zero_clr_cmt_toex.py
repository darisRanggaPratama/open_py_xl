import openpyxl as pyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Buka file Excel: *.xlsx
path = input('Zero Values in Worksheet\nFile Name: ')
workbook = pyxl.load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Warna latar belakang
colored = PatternFill(start_color='F5B7B1', end_color='F5B7B1', fill_type='solid')

# Loop melalui sel di Worksheet target, misal dari A2 sampai F46
for row in worksheet.iter_rows(min_row=2, max_row=46, min_col=1, max_col=6):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            # Periksa apakah nilai Zero
            if cell.value == 0:
                commented = Comment(f"ZERO\n{cell.coordinate}: {cell.value}", "Author")
                cell.comment = None
                cell.comment = commented
                cell.fill = colored
                # Tambahkan hasil ke list
                result.append((cell.coordinate, cell.value))

# Buat worksheet "ZERO" jika belum ada
if 'ZER' not in workbook.sheetnames:
    workbook.create_sheet('ZER')

# Pilih worksheet "ZERO"
sheetZer = workbook['ZER']

# Judul: ZERO
sheetZer.cell(row=1, column=2, value='No; Cell; ZERO')

# Tulis hasil ke dalam Worksheet "ZERO" di cell B2
for i, (coordinate, value) in enumerate(result):
    sheetZer.cell(row=i + 2, column=2, value=f"{i + 1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file Excel: *.xlsx
workbook.save('sample.xlsx')
