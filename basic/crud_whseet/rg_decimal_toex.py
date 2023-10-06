import openpyxl as pyxl

# Buka file Excel: *.xlsx
path = input('Decimal Values in Worksheet\nFile Name: ')
workbook = pyxl.load_workbook(path)

# Pilih worksheet tertentu
print(f'Available Worksheet:\n{workbook.sheetnames}')
sheet = input('Get 1 Worksheet: ')
worksheet = workbook[sheet]

# Inisialisasi list untuk menyimpan hasil
result = []

# Loop melalui sel di Worksheet target, misal dari A2 sampai F46
for row in worksheet.iter_rows(min_row=2, max_row=46, min_col=1, max_col=6):
    for cell in row:
        # Periksa apakah nilai sel adalah Decimal/Float/Double
        if isinstance(cell.value, (float)):
            result.append((cell.coordinate, cell.value))

# Buat worksheet "Hasil" jika belum ada
if 'Hasil' not in workbook.sheetnames:
    workbook.create_sheet('Hasil')

# Pilih worksheet "Hasil"
worksheet_hasil = workbook['Hasil']

# Judul: Desimal
worksheet_hasil.cell(row=1, column=2, value='No; Cell; DESIMAL')

# Tulis hasil ke dalam Worksheet "Hasil" di cell B2
for i, (coordinate, value) in enumerate(result):
    worksheet_hasil.cell(row=i + 2, column=2, value=f"{i+1}; {coordinate}; {value}")

# Simpan perubahan ke dalam file.xlsx
workbook.save('sample.xlsx')
